"""
PDF to Excel conversion service.
"""

import io
import re
import uuid
from datetime import datetime, timedelta
from typing import List, Dict, Tuple, Optional, Any

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.utils import timezone
from .models import TempAnalysis


class PDFToExcelService:
    """Service for converting PDF tables to Excel."""

    def __init__(self):
        self.session_timeout = 10  # minutes

    def analyze_pdf(self, pdf_file: io.BytesIO) -> Dict[str, Any]:
        """
        Analyze PDF and extract table data.
        
        Args:
            pdf_file: PDF file-like object
            
        Returns:
            Dictionary with analysis results
        """
        text = self._extract_text(pdf_file)
        
        if not text or len(text.strip()) == 0:
            raise ValueError("No text found in PDF")
        
        # Extract table data
        columns, rows = self._extract_table_from_text(text)
        
        # Generate analysis ID
        analysis_id = f"analysis_{uuid.uuid4().hex[:12]}"
        
        # Store in database
        expires_at = timezone.now() + timedelta(minutes=self.session_timeout)
        TempAnalysis.objects.create(
            id=analysis_id,
            columns=columns,
            rows=rows,
            expires_at=expires_at
        )
        
        return {
            'analysis_id': analysis_id,
            'columns': columns,
            'sample_rows': rows[:5],
            'total_rows': len(rows)
        }

    def generate_excel(self, analysis_id: str, selected_columns: List[str]) -> Tuple[bytes, str, dict]:
        """
        Generate Excel file from analyzed data.
        
        Args:
            analysis_id: ID from analyze_pdf
            selected_columns: List of column names to include
            
        Returns:
            Tuple of (excel_bytes, filename, stats)
        """
        try:
            analysis = TempAnalysis.objects.get(id=analysis_id)
        except TempAnalysis.DoesNotExist:
            raise ValueError("Session expired. Please upload PDF again.")
        
        columns = analysis.columns
        all_rows = analysis.rows
        
        # Filter columns
        selected_indices = [columns.index(col) for col in selected_columns if col in columns]
        filtered_rows = [
            [row[idx] if idx < len(row) else "" for idx in selected_indices]
            for row in all_rows
        ]
        
        # Create Excel
        excel_bytes = self._create_excel(selected_columns, filtered_rows)
        
        # Clean up
        analysis.delete()
        
        filename = f"converted_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        stats = {
            'total_rows': len(filtered_rows),
            'columns': len(selected_columns),
            'filename': filename
        }
        
        return excel_bytes, filename, stats

    def _extract_text(self, pdf_file: io.BytesIO) -> str:
        """Extract text from PDF using pdfplumber."""
        text = ""
        with pdfplumber.open(pdf_file) as pdf:
            for page in pdf.pages:
                # Try to extract tables first
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        for row in table:
                            text += "\t".join([str(cell) if cell else "" for cell in row]) + "\n"
                else:
                    # Fall back to text extraction
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
        
        return text

    def _extract_table_from_text(self, text: str) -> Tuple[List[str], List[List[str]]]:
        """Extract structured table data from text."""
        lines = text.split("\n")
        lines = [line.strip() for line in lines if len(line.strip()) > 5]
        
        # Try tab-separated first
        if any("\t" in line for line in lines):
            return self._extract_tsv(lines)
        
        # Try multiple space separator
        structured_rows = [line for line in lines if re.search(r'\s{2,}', line)]
        if len(structured_rows) >= 2:
            return self._extract_space_separated(structured_rows)
        
        # Try smart split for contact-like data
        return self._extract_smart_split(lines)

    def _extract_tsv(self, lines: List[str]) -> Tuple[List[str], List[List[str]]]:
        """Extract tab-separated data."""
        rows = []
        for line in lines:
            if "\t" in line:
                cells = [cell.strip() for cell in line.split("\t") if cell.strip()]
                if cells:
                    rows.append(cells)
        
        if not rows:
            raise ValueError("No table data found")
        
        # Find most common column count
        column_counts = {}
        for row in rows:
            count = len(row)
            column_counts[count] = column_counts.get(count, 0) + 1
        
        most_common = max(column_counts.keys(), key=lambda k: column_counts[k])
        valid_rows = [row for row in rows if len(row) == most_common]
        
        return valid_rows[0], valid_rows[1:]

    def _extract_space_separated(self, lines: List[str]) -> Tuple[List[str], List[List[str]]]:
        """Extract data separated by multiple spaces."""
        rows = []
        for line in lines:
            cells = [cell.strip() for cell in re.split(r'\s{2,}', line) if cell.strip()]
            if cells:
                rows.append(cells)
        
        if not rows:
            raise ValueError("No table data found")
        
        # Find most common column count
        column_counts = {}
        for row in rows:
            count = len(row)
            column_counts[count] = column_counts.get(count, 0) + 1
        
        most_common = max(column_counts.keys(), key=lambda k: column_counts[k])
        valid_rows = [row for row in rows if len(row) == most_common]
        
        return valid_rows[0], valid_rows[1:]

    def _extract_smart_split(self, lines: List[str]) -> Tuple[List[str], List[List[str]]]:
        """Smart split for mashed-together contact data."""
        rows = []
        
        for line in lines:
            if len(line) < 20:
                continue
            
            cells = self._smart_split_line(line)
            if cells:
                rows.append(cells)
        
        if not rows:
            raise ValueError("Could not extract data from PDF")
        
        # Determine headers
        first_row = rows[0]
        if len(first_row) == 3 and any(c.isdigit() for c in first_row[2]):
            columns = ["Name", "Address", "Phone"]
        else:
            columns = [f"Column {i+1}" for i in range(len(first_row))]
        
        return columns, rows

    def _smart_split_line(self, line: str) -> Optional[List[str]]:
        """Split a single line intelligently."""
        # Try to find phone number pattern
        phone_pattern = r'(\+?\d{1,4}[-\s]?\d{3,4}[-\s]?\d{4,})'
        phone_match = re.search(phone_pattern, line)
        
        if phone_match:
            phone = phone_match.group(0).strip()
            before_phone = line[:line.index(phone)].strip()
            
            # Try to split name and address
            first_digit = next((i for i, c in enumerate(before_phone) if c.isdigit()), -1)
            
            if first_digit > 0:
                name = before_phone[:first_digit].strip()
                address = before_phone[first_digit:].strip()
                return [name, address, phone]
            
            # Try multiple spaces
            if "  " in before_phone:
                parts = [p.strip() for p in before_phone.split("  ", 1)]
                return [parts[0], parts[1] if len(parts) > 1 else "", phone]
            
            return [before_phone, "", phone]
        
        # Try tab or multiple space split
        if "\t" in line or "   " in line:
            sep = "\t" if "\t" in line else "   "
            cells = [c.strip() for c in line.split(sep) if c.strip()]
            if len(cells) >= 2:
                return cells
        
        # Try comma split
        if line.count(",") >= 2:
            cells = [c.strip() for c in line.split(",") if c.strip()]
            if len(cells) >= 3:
                return cells
        
        return None

    def _create_excel(self, headers: List[str], rows: List[List[str]]) -> bytes:
        """Create formatted Excel file."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # Header styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        ws.row_dimensions[1].height = 25
        
        # Add data
        alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
        
        # Auto-size columns
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in rows:
                if col_idx <= len(row):
                    max_length = max(max_length, len(str(row[col_idx - 1])))
            
            adjusted_width = min(max(max_length + 2, 12), 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.read()

    def cleanup_expired(self):
        """Clean up expired analysis sessions."""
        TempAnalysis.objects.filter(expires_at__lt=timezone.now()).delete()


# Singleton instance
pdf_to_excel_service = PDFToExcelService()
